"""Tests for Phase 9 — Report generation (CSV, Excel, PDF)."""
import pytest
import pytest_asyncio
from mongomock_motor import AsyncMongoMockClient

from app.services.report_service import ReportService


@pytest_asyncio.fixture
async def db():
    client = AsyncMongoMockClient()
    d = client["test_reports"]
    await d["sessions"].insert_many([
        {"session_id": "S1", "camera_id": "CAM01", "type": "LOADING",
         "status": "COMPLETED", "start_time": "2026-06-17T10:00:00Z",
         "end_time": "2026-06-17T10:30:00Z", "count": 10},
        {"session_id": "S2", "camera_id": "CAM02", "type": "UNLOADING",
         "status": "ACTIVE", "start_time": "2026-06-17T11:00:00Z",
         "end_time": None, "count": 5},
    ])
    await d["count_events"].insert_many([
        {"event_id": "e1", "camera_id": "CAM01", "session_id": "S1",
         "track_id": 1, "direction": "LOADING", "timestamp": "2026-06-17T10:05:00Z"},
        {"event_id": "e2", "camera_id": "CAM01", "session_id": "S1",
         "track_id": 2, "direction": "LOADING", "timestamp": "2026-06-17T10:06:00Z"},
        {"event_id": "e3", "camera_id": "CAM02", "session_id": "S2",
         "track_id": 3, "direction": "UNLOADING", "timestamp": "2026-06-17T11:05:00Z"},
    ])
    return d


@pytest_asyncio.fixture
async def svc(db):
    return ReportService(db)


class TestCSV:
    @pytest.mark.asyncio
    async def test_csv_generates_bytes(self, svc):
        data = await svc.generate_csv()
        assert isinstance(data, bytes)
        text = data.decode("utf-8")
        assert "session_id" in text
        assert "S1" in text
        assert "e1" in text

    @pytest.mark.asyncio
    async def test_csv_filters_by_camera(self, svc):
        data = await svc.generate_csv(camera_id="CAM01")
        text = data.decode("utf-8")
        assert "CAM01" in text
        assert "CAM02" not in text


class TestExcel:
    @pytest.mark.asyncio
    async def test_excel_generates_bytes(self, svc):
        data = await svc.generate_excel()
        assert isinstance(data, bytes)
        # Excel files start with PK (ZIP format)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_excel_has_sheets(self, svc):
        from openpyxl import load_workbook
        import io
        data = await svc.generate_excel()
        wb = load_workbook(io.BytesIO(data))
        assert "Sessions" in wb.sheetnames
        assert "Count Events" in wb.sheetnames
        ws = wb["Sessions"]
        assert ws.cell(1, 1).value == "session_id"
        assert ws.max_row >= 3  # header + 2 sessions


class TestPDF:
    @pytest.mark.asyncio
    async def test_pdf_generates_bytes(self, svc):
        data = await svc.generate_pdf()
        assert isinstance(data, bytes)
        assert data[:5] == b"%PDF-"

    @pytest.mark.asyncio
    async def test_pdf_filters_by_date(self, svc):
        data = await svc.generate_pdf(
            start_date="2026-06-17T10:00:00Z",
            end_date="2026-06-17T10:59:00Z",
        )
        assert len(data) > 0
